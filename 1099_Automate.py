#to do: finish ironing out the address mapper (create mapping dictionary)


from PyPDF2 import PdfFileWriter, PdfFileReader, PdfFileMerger,PdfMerger
import io, openpyxl, random
from openpyxl import load_workbook


#Jan 11th coding sesh:
#Organize excel data into nested dictionary like below:
#{'Victor': {'llc': 'westwind', 'payer_TIN': '12345', 'recipient_TIN': '12345'},
#'Brian': {'llc': 'holiday', 'payer_TIN': '12345', 'recipient_TIN': '98754'}}

def llc(parkname):
    d = {'Hitching Post':'Hitching Post Mobile Home Park, LLC',
    'Crestview':'Yucaipa Crestview, LLC',
    'Westwind':'Yucaipa Westwind Estates, LLC',
    'Holiday':'Holiday Rancho Park, LLC',
    'Wishing Well':'Wishing Well Mobile Home Park, LLC',
    'Patrician':'Patrician Mobile Home Park',
    'Mt Vista':'Mount Vista, LLC',
    'Jian Personal':'Jian Chen',
    'Banning':'Banning Wilson Gardens, LLC'
    }
    end_of_string = """11034 Deer Canyon Dr, \nRancho Cucamonga, CA 91737 \n909-210-1491"""
    for i in d:
        if i in parkname:
            return d[i]+ """\n"""+ end_of_string
    return 'N/A'

#takes inputs from excel input file, organizes it into dictionary d
def openpyxl():
    path = r'C:\Users\Lenovo\PycharmProjects\1099_automating\input_template.xlsx'
    file = load_workbook(path)
    sheet = file.active

    #Jan 11th coding (starting fresh)
    #'Brian': {'Address': 'a', 'Payer_TIN': 'b', 'Recipient_TIN': 'c','Name':'Brian','Street':'d'}
    d = {}

    #for x in all excel rows:
    for row in range(sheet.max_row):
        if sheet.cell(row=row+4,column=3).value != None:
            Name = sheet.cell(row=row+4,column=3).value
            Park_Name = sheet.cell(row=row+4,column=4).value
            Payer_TIN = sheet.cell(row=row+4,column=5).value
            Recipient_TIN = sheet.cell(row=row+4,column=6).value
            Street = sheet.cell(row=row+4,column=7).value
            City_state_zip = sheet.cell(row=row+4,column=8).value
            Compensation = sheet.cell(row=row+4,column=9).value
        
##        Name = 'Victor'
##        Address = 'a'
##        Payer_TIN = 'b'
##        Recipient_TIN = 'c'
##        Street = 'd'
##        City_state_zip = 'e'
##        Compensation = 'f'

        values = {'Name':Name,'Address':llc(Park_Name),'Payer_TIN':Payer_TIN,'Recipient_TIN':Recipient_TIN,
                  'Street':Street,'City_state_zip':City_state_zip,'Compensation':Compensation}
        d[Name] = values
        
##    d = {}
##    for row in range(sheet.max_row):
##        num = sheet.cell(row=row+1, column=2).value
##        field = sheet.cell(row=row+1, column=3).value
##        if type(num)==int  and field !=None:
##            raw_input = sheet.cell(row=row + 1, column=4).value
##            d[field] = raw_input
    print(d)
    return d

openpyxl()

#given a park name, return the corresponding LLC
#format will be in multi row format, as follows:
#Banning Wilson Gardens, LLC
#11034 Deer Canyon Dr,
#Rancho Cucamonga, CA 91737
#909-210-1491
#Holiday, Westwind,Banning,Crestview,Mt Vista,Jian Personal,Wishing Well,Hitching Post


def alterpdf():
##    d = {'Victor': {'Address': 'a', 'Payer_TIN': 'b', 'Recipient_TIN': 'c','Name':'Victor','Street':'d'},
##         'Brian': {'Address': 'a', 'Payer_TIN': 'b', 'Recipient_TIN': 'c','Name':'Brian','Street':'d'}}
    d = openpyxl()
    emptypath = 'C:\\Users\\Lenovo\\PycharmProjects\\1099_automating\\empty_template.pdf'
    filledpath = 'C:\\Users\\Lenovo\\PycharmProjects\\1099_automating\\output\\'
    reader = PdfFileReader(emptypath)
    writer = PdfFileWriter()
    page = reader.pages[0]
    page1 = reader.pages[1]
    page2 = reader.pages[2]
    page3 = reader.pages[3]
    page4 = reader.pages[4]
    
    fields = reader.getFields()
    writer.addPage(page)
    writer.addPage(page1)
    writer.addPage(page2)
    writer.addPage(page3)
    writer.addPage(page4)
    
    #I will butcher this
    for person in d:
         for x in d[person]:
             print('we should now be updating the "person" pdf')
             writer.updatePageFormFieldValues(writer.getPage(0), {x: d[person][x]})
             writer.updatePageFormFieldValues(writer.getPage(1),{x: d[person][x]})
             writer.updatePageFormFieldValues(writer.getPage(2),{x: d[person][x]})
             writer.updatePageFormFieldValues(writer.getPage(3),{x: d[person][x]})   
             writer.updatePageFormFieldValues(writer.getPage(4),{x: d[person][x]})
             #write different pdf output for each person
             
             output_file_name = filledpath+ person+'.pdf'
             with open(output_file_name, "wb") as output_stream:
                writer.write(output_stream)
    return None
##alterpdf()

#Jan 17th: I want only the third page to print out for contractors
def extract_third_page():
    d = openpyxl()
    original_path = 'C:\\Users\\Lenovo\\PycharmProjects\\1099_automating\\output\\'
    new_path = 'C:\\Users\\Lenovo\\PycharmProjects\\1099_automating\\output2\\'
    for person in d:
        pdfname = original_path+person+'.pdf'
        inputpdf = PdfFileReader(open(pdfname, "rb"))
        for i in range(inputpdf.numPages):
            output = PdfFileWriter()
            output.addPage(inputpdf.getPage(2))
            output_file_name = new_path+person+".pdf"
            with open(output_file_name, "wb") as outputStream:
                output.write(outputStream)

extract_third_page()

#below code is no good because of adobe's damned form fields
#Jan 17 2023 code - combine third page of PDFs into single folder
#merge only the third page of each PDF
def merge():
    d = openpyxl()
    path = 'C:\\Users\\Lenovo\\PycharmProjects\\1099_automating\\output\\'
    pdfs = []
    for person in d:
        pdfs.append(path+person+".pdf")
##    pdfs = ['file1.pdf', 'file2.pdf', 'file3.pdf', 'file4.pdf']
##    print(pdfs)
    merger = PdfMerger()

    print(pdfs)
    for pdf in pdfs:
        merger.append(pdf)
##        merger.append(pdf, pages=(2,3))

    merger.write('C:\\Users\\Lenovo\\PycharmProjects\\1099_automating\\combined.pdf')
    merger.close()
    return None
##merge()
##    # Now you add your data to the forms!
##    for x in d:
##        writer.updatePageFormFieldValues(
##            writer.getPage(0), {x: d[x]}
##        )
##    # write "output" to PyPDF2-output.pdf
##    with open(filledpath, "wb") as output_stream:
##        writer.write(output_stream)
###fill up them PDFs baby
##def fill():
##    L = ['page1', 'page2', 'page3', 'page4', \
##         'page5','page6','page7','page8','page9']
##    for i in L:
##        emptypath = 'C:\\Users\\Lenovo\\PycharmProjects\\YucaipaCityPermit\\input\\'+ i +'.pdf'
##        filledpath = 'C:\\Users\\Lenovo\\PycharmProjects\\YucaipaCityPermit\\output\\' + i +'.pdf'
##        alterpdf(emptypath,filledpath)
##
###combine every file in the filled path
##def combine():
##    merger = PdfFileMerger()
##    L = ['page1', 'page2', 'page3', 'page4', \
##         'page5', 'page6', 'page7','page8','page9','drawingoutput','drawingoutput','drawingoutput']
##    for i in L:
##        file = 'C:\\Users\\Lenovo\\PycharmProjects\\YucaipaCityPermit\\output\\' + i +'.pdf'
##        merger.append(PdfFileReader(open(file,'rb')))
##    merger.write(r'C:\Users\Lenovo\PycharmProjects\YucaipaCityPermit\printme\combined.pdf')
##    return None

